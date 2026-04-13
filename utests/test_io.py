import json
from unittest.mock import MagicMock, mock_open, patch

import pytest
from openpyxl import load_workbook

from pytabify import DataTableCreator, DataTableSaver
from pytabify.adapters.files.errors import (
    FileExtensionException,
    FileNotFoundException,
    FileReadingException,
    FileWritingException,
    SheetNameDoesNotExistException,
    SheetNameHasNotEmptyException,
)
from pytabify.adapters.files.readers import (
    CSVFileReadingAdapter,
    JSONFileReadingAdapter,
    XLSXReadingAdapter,
)
from pytabify.adapters.files.writers import (
    CsvFileWritingAdapter,
    JsonFileWritingAdapter,
    XlsxFileWritingAdapter,
)


@pytest.fixture
def sample_datatable():
    return DataTableCreator.from_records(
        [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25},
        ]
    )


def test_from_file_json_end_to_end(tmp_path):
    source = [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]
    input_file = tmp_path / "input.json"
    input_file.write_text(json.dumps(source), encoding="utf-8")

    datatable = DataTableCreator.from_file(str(input_file))

    assert datatable.to_dict() == source


def test_from_file_csv_end_to_end(tmp_path):
    input_file = tmp_path / "input.csv"
    input_file.write_text("name,age\nAlice,30\nBob,25\n", encoding="utf-8")

    datatable = DataTableCreator.from_file(str(input_file))

    assert datatable.to_dict() == [{"name": "Alice", "age": "30"}, {"name": "Bob", "age": "25"}]


def test_from_file_invalid_extension():
    with pytest.raises(FileExtensionException):
        DataTableCreator.from_file("archivo.txt")


@patch("pathlib.Path.open", new_callable=mock_open, read_data='[{"name": "Alice"}]')
def test_json_reading_adapter(mock_file, tmp_path):
    filepath = tmp_path / "test.json"
    filepath.write_text('[{"name": "Alice"}]')
    data = JSONFileReadingAdapter().read(str(filepath))
    assert data[0]["name"] == "Alice"


@patch("pathlib.Path.open", new_callable=mock_open, read_data="name,age\nAlice,30\nBob,25\n")
def test_csv_reading_adapter(mock_file, tmp_path):
    filepath = tmp_path / "test.csv"
    filepath.write_text("name,age\nAlice,30\nBob,25\n")
    data = CSVFileReadingAdapter().read(str(filepath))
    assert data[1]["age"] == "25"


def test_csv_reading_adapter_oserror_raises_reading_error(tmp_path):
    filepath = tmp_path / "test.csv"
    filepath.write_text("name,age\nAlice,30\n", encoding="utf-8")

    with (
        patch("pathlib.Path.open", side_effect=OSError("boom")),
        pytest.raises(FileReadingException, match="abrir el archivo de datos csv"),
    ):
        CSVFileReadingAdapter().read(str(filepath))


@patch("pytabify.adapters.files.readers.xlsx_file_reading_adapter.load_workbook")
def test_xlsx_reading_adapter(mock_load_workbook, tmp_path):
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_workbook.__getitem__.return_value = mock_sheet
    mock_workbook.sheetnames = ["Sheet"]
    mock_load_workbook.return_value = mock_workbook
    mock_sheet.__getitem__.return_value = [MagicMock(value="name"), MagicMock(value="age")]
    mock_sheet.iter_rows.return_value = [
        ("Alice", 30),
        ("Bob", 25),
    ]

    filepath = tmp_path / "test.xlsx"
    filepath.touch()

    data = XLSXReadingAdapter().read(str(filepath), sheet_name="Sheet")

    assert data == [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]


@pytest.mark.parametrize(
    ("reader_class", "path"),
    [
        (JSONFileReadingAdapter, "test.json"),
        (CSVFileReadingAdapter, "test.csv"),
        (XLSXReadingAdapter, "test.xlsx"),
    ],
)
def test_readers_raise_consistent_file_not_found(reader_class, path):
    kwargs = {"sheet_name": "Sheet1"} if reader_class is XLSXReadingAdapter else {}
    with pytest.raises(FileNotFoundException):
        if reader_class is XLSXReadingAdapter:
            reader_class().read(path, **kwargs)
        else:
            reader_class().read(path, **kwargs)


@patch("pathlib.Path.open", new_callable=mock_open, read_data="{bad json}")
def test_json_invalid_content_raises_reading_error(mock_file, tmp_path):
    filepath = tmp_path / "test.json"
    filepath.write_text("{bad json}")
    with pytest.raises(FileReadingException):
        JSONFileReadingAdapter().read(str(filepath))


@patch("pytabify.adapters.files.readers.xlsx_file_reading_adapter.load_workbook")
def test_xlsx_reading_adapter_sheet_name_does_not_exist(mock_load_workbook, tmp_path):
    mock_workbook = MagicMock()
    mock_workbook.sheetnames = ["Sheet1", "Sheet2"]
    mock_load_workbook.return_value = mock_workbook

    filepath = tmp_path / "test.xlsx"
    filepath.touch()

    with pytest.raises(SheetNameDoesNotExistException):
        XLSXReadingAdapter().read(str(filepath), sheet_name="SheetTest")


@patch("pytabify.adapters.files.readers.xlsx_file_reading_adapter.load_workbook")
def test_xlsx_reading_adapter_sheet_name_none(mock_load_workbook, tmp_path):
    mock_workbook = MagicMock()
    mock_load_workbook.return_value = mock_workbook

    filepath = tmp_path / "test.xlsx"
    filepath.touch()

    with pytest.raises(SheetNameHasNotEmptyException):
        XLSXReadingAdapter().read(str(filepath))


@patch("pathlib.Path.open", new_callable=mock_open)
def test_into_json(mock_file, sample_datatable, tmp_path):
    output_file = tmp_path / "output.json"
    DataTableSaver.into_json(sample_datatable, str(output_file))
    handle = mock_file()
    handle.write.assert_called()


@patch("pathlib.Path.open", new_callable=mock_open)
def test_into_csv(mock_file, sample_datatable, tmp_path):
    output_file = tmp_path / "output.csv"
    DataTableSaver.into_csv(sample_datatable, str(output_file))
    handle = mock_file()
    handle.write.assert_called()


def test_into_csv_end_to_end(tmp_path, sample_datatable):
    output_file = tmp_path / "output.csv"

    DataTableSaver.into_csv(sample_datatable, str(output_file))

    assert output_file.read_text(encoding="utf-8").splitlines() == [
        "name,age",
        "Alice,30",
        "Bob,25",
    ]


def test_into_xlsx_end_to_end(tmp_path, sample_datatable):
    from openpyxl import load_workbook

    output_file = tmp_path / "output.xlsx"

    DataTableSaver.into_xlsx(sample_datatable, str(output_file))

    workbook = load_workbook(output_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows == [("name", "age"), ("Alice", 30), ("Bob", 25)]


@patch(
    "pytabify.adapters.files.writers.xlsx_file_writing_adapter.Workbook.save",
    side_effect=Exception("Error"),
)
def test_into_xlsx_error(mock_save, sample_datatable, tmp_path):
    with pytest.raises(FileWritingException):
        XlsxFileWritingAdapter().write(sample_datatable, str(tmp_path / "output.xlsx"))


def test_into_csv_error(sample_datatable, tmp_path):
    with (
        patch("csv.DictWriter.writerow", side_effect=Exception("Error")),
        pytest.raises(FileWritingException),
    ):
        CsvFileWritingAdapter().write(sample_datatable, str(tmp_path / "output.csv"), "utf-8")


def test_into_json_error(sample_datatable, tmp_path):
    with (
        patch("json.dump", side_effect=Exception("Error")),
        pytest.raises(FileWritingException),
    ):
        JsonFileWritingAdapter().write(sample_datatable, str(tmp_path / "output.json"), "utf-8")


def test_json_round_trip(tmp_path):
    source = [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]
    datatable = DataTableCreator.from_records(source)
    output_file = tmp_path / "output.json"

    DataTableSaver.into_json(datatable, str(output_file))

    saved_data = json.loads(output_file.read_text(encoding="utf-8"))
    assert saved_data == source


def test_xlsx_round_trip(tmp_path):
    from openpyxl import Workbook

    from pytabify.adapters.files.readers import xlsx_file_reading_adapter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "People"
    sheet.append(["name", "age"])
    sheet.append(["Alice", 30])
    sheet.append(["Bob", 25])

    input_file = tmp_path / "input.xlsx"
    workbook.save(input_file)
    xlsx_file_reading_adapter.load_workbook = load_workbook

    datatable = DataTableCreator.from_file(str(input_file), sheet_name="People")

    assert datatable.to_dict() == [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]
