from openpyxl import load_workbook

from pytabify.adapters.files.readers import xlsx_file_reading_adapter
from pytabify.io.interfaces.read import ReadingStrategy
from pytabify.adapters.files.readers import (
    CSVFileReadingAdapter,
    JSONFileReadingAdapter,
    XLSXReadingAdapter,
)

class JSONFileReadingStrategy(ReadingStrategy):
    """JsonFileReadingStrategy"""
    def read(self):
        return JSONFileReadingAdapter().read(self._path, **self._kwargs)

class CSVFileReadingStrategy(ReadingStrategy):
    """CsvFileReadingStrategy"""
    def read(self):
        return CSVFileReadingAdapter().read(self._path, **self._kwargs)

class XLSXReadingStrategy(ReadingStrategy):
    """XlsxReadingStrategy"""
    def read(self):
        xlsx_file_reading_adapter.load_workbook = load_workbook
        return XLSXReadingAdapter().read(self._path, **self._kwargs)
