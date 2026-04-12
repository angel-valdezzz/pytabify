from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from pytabify.application.ports.table_reader import TableReader
from pytabify.utils.errors import (
    FileNotFoundException,
    FileReadingException,
    SheetNameDoesNotExistException,
    SheetNameHasNotEmptyException,
)


class XLSXReadingAdapter(TableReader):
    def read(self, path: str, **kwargs) -> list[dict[str, Any]]:
        sheet_name = kwargs.get("sheet_name")
        if sheet_name is None:
            raise SheetNameHasNotEmptyException("sheet_name debe ser definido")

        try:
            workbook = load_workbook(path)
        except FileNotFoundError as exc:
            raise FileNotFoundException(f"El archivo {path} NO Existe verifique la ruta.") from exc
        except OSError as exc:
            raise FileReadingException("Ocurrio un error al abrir el archivo de datos xlsx") from exc

        if sheet_name not in workbook.sheetnames:
            raise SheetNameDoesNotExistException(f"La hoja {sheet_name} no existe en el archivo")

        sheet = workbook[sheet_name]
        headers = [str(cell.value) for cell in sheet[1]]
        data: list[dict[str, Any]] = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for index, cell_value in enumerate(row):
                row_data[headers[index]] = cell_value
            data.append(row_data)

        return data
