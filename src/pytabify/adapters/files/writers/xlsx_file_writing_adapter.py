from openpyxl import Workbook

from pytabify.application.ports.table_writer import TableWriter
from pytabify.domain.data_table import DataTable
from pytabify.utils.errors import FileWritingException


class XlsxFileWritingAdapter(TableWriter):
    def write(self, datatable: DataTable, path: str, encoding: str = "utf-8"):
        del encoding

        fieldnames = [header.name for header in datatable.headers()]
        data = datatable.to_dict()
        workbook = Workbook()
        sheet = workbook.active

        for column_index, header in enumerate(fieldnames, start=1):
            sheet.cell(row=1, column=column_index, value=header)

        for row_index, row_data in enumerate(data, start=2):
            for column_index, header in enumerate(fieldnames, start=1):
                sheet.cell(row=row_index, column=column_index, value=row_data.get(header, ""))

        try:
            workbook.save(path)
        except Exception as exc:
            raise FileWritingException(
                f"No fue posible guardar los datos en el xlsx {path}. Mas detalles: {exc}"
            ) from exc
